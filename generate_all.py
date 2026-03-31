#!/usr/bin/env python3
"""
generate_all.py  —  CONC Kitchen production pipeline generator
Reads: conc_production_data.csv  (the single source of truth)
Reads: hub_template parts (shell HTML, RCP block, logic JS) extracted from last hub build
Writes: CONC_Production_Hub.html
        01_Production_Schedule.xlsx
        03_Driver_Schedule.xlsx
        Labour_Report.xlsx

Usage:
  python3 generate_all.py              # generate all outputs
  python3 generate_all.py --hub-only   # hub only
  python3 generate_all.py --xlsx-only  # xlsx outputs only
"""

import csv, re, sys, json, subprocess, textwrap
from pathlib import Path
from collections import defaultdict

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE         = Path(__file__).parent
XLSX_PATH    = BASE / '00_Production_Source_Table.xlsx'   # primary source
CSV_PATH     = BASE / 'conc_production_data.csv'          # fallback
SHELL_PATH   = BASE / 'hub_shell.html'
RCP_PATH     = BASE / 'hub_rcp.js'
LOGIC_PATH   = BASE / 'hub_logic.js'
HUB_TEMPLATE = BASE / 'CONC_Production_Hub.html'          # carries FRIDGE + MOVES
OUTPUT_DIR   = BASE / 'outputs'

# ── Date / Week config ─────────────────────────────────────────────────────────
DAYS     = ['SUNDAY','MONDAY','TUESDAY','WEDNESDAY','THURSDAY','FRIDAY','SATURDAY']
DAY_NAMES = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']

# (week, day_index 0=Sun) → (display_str, dateNum)
DATE_MAP = {
    (1,0):('May 17',17),(1,1):('May 18',18),(1,2):('May 19',19),(1,3):('May 20',20),
    (1,4):('May 21',21),(1,5):('May 22',22),(1,6):('May 23',23),
    (2,0):('May 24',24),(2,1):('May 25',25),(2,2):('May 26',26),(2,3):('May 27',27),
    (2,4):('May 28',28),(2,5):('May 29',29),(2,6):('May 30',30),
    (3,0):('May 31',31),(3,1):('Jun 1',1),(3,2):('Jun 2',2),(3,3):('Jun 3',3),
    (3,4):('Jun 4',4),(3,5):('Jun 5',5),(3,6):('Jun 6',6),
    (4,0):('Jun 7',7),(4,1):('Jun 8',8),(4,2):('Jun 9',9),(4,3):('Jun 10',10),
    (4,4):('Jun 11',11),(4,5):('Jun 12',12),(4,6):('Jun 13',13),
}
WEEK_RANGES = {
    1:'May 17–23, 2026', 2:'May 24–30, 2026',
    3:'May 31–Jun 6, 2026', 4:'Jun 7–13, 2026'
}
WEEK_NOTE = 'Remediation · All food HOT from Bloor or LAN · Rex = portion & serve only'

SECTION_ORDER  = ['SEND AM','LUNCH','PRODUCTION','DINNER','SEND PM']
SECTION_IDS    = {'SEND AM':'send-am','LUNCH':'lunch','PRODUCTION':'production',
                  'DINNER':'dinner','SEND PM':'send-pm'}
SECTION_LABELS = {'send-am':'▲ Send AM','lunch':'☀ Lunch','production':'⚙ Production',
                  'dinner':'🌙 Dinner','send-pm':'▼ Send PM'}

# ── Labour rules ───────────────────────────────────────────────────────────────
LONG_COOK_ITEMS = [
    'massaman','arroz con pollo','fried rice','white chili','white chilli',
    'sous vide','sausage pasta sauce',
    'west african peanut stew','caribbean stew','beef & vegetable stew',
    'vegan chili','chickpea shakshuka','beef stroganoff'
]
OVERNIGHT_ITEMS = ['philly steak overnight']
def labour_times(row):
    """Derive active_min and oven_min from cook_min + item name + type."""
    try:
        cook_min = int(str(row.get('cook_min') or '0').replace(' min','').strip())
    except (ValueError, TypeError):
        cook_min = 0
    typ = (row.get('type') or '').upper()
    item_lower = (row.get('item') or '').lower()
    if typ == 'HEAT':
        return 15, 0
    if any(k in item_lower for k in OVERNIGHT_ITEMS):
        return 30, cook_min
    if any(k in item_lower for k in LONG_COOK_ITEMS):
        return cook_min // 2, cook_min
    return cook_min, cook_min

# ── Stream detection ───────────────────────────────────────────────────────────
def detect_stream(item):
    item_l = item.lower()
    if 'vegan' in item_l or '(veg)' in item_l or 'vegetarian' in item_l:
        return 'Vegan'
    if 'halal' in item_l:
        return 'Halal'
    return 'Regular'

# ── Source loader (xlsx primary, CSV fallback) ────────────────────────────────
def load_source():
    if XLSX_PATH.exists():
        return _load_xlsx()
    elif CSV_PATH.exists():
        print(f"  (xlsx not found, falling back to CSV)")
        return _load_csv()
    else:
        raise FileNotFoundError(f"No source file found. Expected {XLSX_PATH} or {CSV_PATH}")

def _load_xlsx():
    import openpyxl
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    data_sheets = [s for s in wb.sheetnames if s != 'README']
    headers = None
    rows = []
    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        if headers is None:
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            headers = [h for h in headers if h]  # drop None trailing cols
        for r in range(2, ws.max_row + 1):
            item = ws.cell(r, 5).value  # item col
            if not item:
                continue
            row = {}
            for ci, h in enumerate(headers, 1):
                v = ws.cell(r, ci).value
                row[h] = v if v is not None else ''
            row['week'] = int(row['week']) if row['week'] else 0
            row['cook_min'] = str(row.get('cook_min') or '')
            # Derive dest_fridge from hold_days + override
            row['hold_days'] = _safe_int(row.get('hold_days'))
            row['dest_fridge'] = _derive_dest_fridge(row)
            rows.append(row)
    return rows

def _safe_int(val):
    """Convert to int or return 0."""
    try:
        return int(val) if val else 0
    except (ValueError, TypeError):
        return 0

def _derive_dest_fridge(row):
    """Derive destination fridge: override wins, then ≤2 days = Bloor, 3+ = Rex."""
    override = str(row.get('dest_override') or '').strip()
    if override:
        return override
    hold = row.get('hold_days', 0)
    if not hold or hold == 0:
        return ''
    return 'Bloor' if hold <= 2 else 'Rex'

def _load_csv():
    rows = []
    with open(CSV_PATH, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            row['week'] = int(row['week'])
            row['cook_min'] = row.get('cook_min') or ''
            row['hold_days'] = _safe_int(row.get('hold_days'))
            row['dest_fridge'] = _derive_dest_fridge(row)
            rows.append(row)
    return rows

# ══════════════════════════════════════════════════════════════════════════════
#  HUB GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def js_str(s):
    """Escape a string for single-quoted JS."""
    return str(s or '').replace('\\','\\\\').replace("'","\\'")

def build_serves(row, wk, day_idx):
    """Return serves string for WK item, or None.
    Only PRODUCTION items get a serves field — LUNCH/DINNER items are same-day service.
    SEND items get route instead.
    """
    period  = row.get('period','')
    raw     = str(row.get('prod_date_raw','') or '')

    # SEND and same-day service sections: no serves
    if period in ('SEND AM','SEND PM','LUNCH','DINNER'):
        return None

    # PRODUCTION items only
    if raw.startswith('Serves:'):
        return '→ ' + raw.replace('Serves:','').strip()
    if raw.startswith('Produced:'):
        txt = raw.replace('Produced:','').strip()
        if 'same day' in txt.lower():
            return None
        if 'prev. cycle' in txt.lower() or 'prev cycle' in txt.lower():
            return '→ Prev. cycle'
        return '→ ' + txt
    return None

def build_route(row):
    """Extract route string for SEND items."""
    for src in [row.get('notes',''), row.get('item','')]:
        src = str(src or '')
        for pat in ['Bloor → Rex','LAN → Rex','GC → Rex','Rex → Bloor',
                    'Vendor → Rex']:
            if pat in src:
                return pat
        # Also catch variants like "LAN → Rex fridge", "Bloor → Rex fridge"
        m = re.match(r'.*((?:Bloor|LAN|GC|Rex|Vendor)\s*→\s*(?:Bloor|LAN|GC|Rex))', src)
        if m:
            return m.group(1)
    return 'Bloor → Rex'

ROUTE_PATTERNS = [
    'Bloor → Rex', 'LAN → Rex', 'GC → Rex', 'Rex → Bloor', 'Vendor → Rex',
    'Bloor → Rex fridge', 'LAN → Rex fridge',
]
def clean_driver_notes(raw_notes, route):
    """Strip route string from notes, return only useful info for driver schedule."""
    notes = str(raw_notes or '').strip()
    if not notes:
        return ''
    # Strip route prefix (exact or with "fridge" suffix)
    for pat in ROUTE_PATTERNS:
        if notes.startswith(pat):
            notes = notes[len(pat):].strip()
            break
    # If remainder starts with pipe or semicolon separator, take what's after it
    if notes.startswith('|') or notes.startswith(';'):
        notes = notes[1:].strip()
    # If the whole note was just the route, nothing left
    if not notes or notes == route:
        return ''
    return notes

def build_flag(row):
    combined = (str(row.get('notes','') or '') + str(row.get('item','') or '')).lower()
    if '⚠' in combined or 'no recipe' in combined:
        return 'warn'
    return None

def s_call(row, wk, day_idx):
    """Render one S() call."""
    typ   = js_str(row['type'])
    item  = js_str(row['item'])
    qty   = js_str(row.get('qty',''))
    site  = js_str(row.get('site',''))
    notes = str(row.get('notes','') or '')
    cook  = str(row.get('cook_min','') or '')
    period= row.get('period','')

    opts = {}
    if cook:
        opts['time'] = 'Overnight' if cook == '480' else f'{cook} min'
    if period in ('SEND AM','SEND PM'):
        opts['route'] = build_route(row)
        if notes and notes not in (opts.get('route',''),):
            opts['notes'] = notes
    else:
        if notes:
            opts['notes'] = notes
    serves = build_serves(row, wk, day_idx)
    if serves:
        opts['serves'] = serves
    flag = build_flag(row)
    if flag:
        opts['flag'] = flag

    opts_str = ','.join(f"{k}:'{js_str(v)}'" for k,v in opts.items())
    if opts_str:
        return f"S('{typ}','{item}','{qty}','{site}',{{{opts_str}}})"
    return f"S('{typ}','{item}','{qty}','{site}')"

def van_load_sum(section_rows):
    """Compute total hotel equiv for a send section."""
    total = 0.0
    has_data = False
    for r in section_rows:
        try:
            v = float(r.get('hotel_equiv') or 0)
            if v > 0:
                total += v
                has_data = True
        except:
            pass
    if not has_data:
        return f'~{len(section_rows)} items'
    fmt = int(total) if total == int(total) else total
    return f'{fmt} HOTELS (hot)'

def build_wk_data(rows, wk):
    """Build WKn JS array string for one week."""
    # Organise rows by day → section
    by_day = defaultdict(lambda: defaultdict(list))
    for r in rows:
        if r['week'] != wk:
            continue
        day_upper = r['day'].strip().upper()
        if day_upper not in DAYS:
            continue
        by_day[day_upper][r['period']].append(r)

    day_objects = []
    for day_idx, day_upper in enumerate(DAYS):
        date_str, date_num = DATE_MAP[(wk, day_idx)]
        day_name = DAY_NAMES[day_idx]
        sections_js = []
        for period in SECTION_ORDER:
            sec_rows = by_day[day_upper].get(period, [])
            if not sec_rows:
                continue
            sid   = SECTION_IDS[period]
            label = SECTION_LABELS[sid]
            items_js = ',\n'.join(s_call(r, wk, day_idx) for r in sec_rows)
            if period in ('SEND AM','SEND PM'):
                vl = van_load_sum(sec_rows)
                vl_part = f",vanLoad:'{js_str(vl)}'" if vl else ''
                sections_js.append(
                    f"{{id:'{sid}',label:'{label}'{vl_part},items:[{items_js}]}}"
                )
            else:
                sections_js.append(
                    f"{{id:'{sid}',label:'{label}',items:[{items_js}]}}"
                )
        sections_str = ',\n'.join(sections_js)
        day_objects.append(
            f"{{dayName:'{day_name}',date:'{date_str}',dateNum:{date_num},"
            f"sections:[{sections_str}]}}"
        )

    days_str = ',\n'.join(day_objects)
    return f"const WK{wk}=[{days_str}];"

def build_meals(rows):
    """Build MEALS JS object from LUNCH/DINNER item names."""
    meals = {}  # dateNum → {lunch, dinner}

    def extract_meal_name(items):
        """Get clean meal summary from a list of rows."""
        names = []
        for r in items:
            item = r.get('item','')
            # Strip trailing parentheticals like (Mon Dinner — Veg), → labels, etc.
            name = re.sub(r'\s*\(.*?\)', '', item)
            name = re.sub(r'\s*→.*$', '', name)
            name = re.sub(r'^(Cook|Heat|Reheat|Steam|Roast|Bake)\s+', '', name, flags=re.I)
            name = name.strip()
            if name and name not in names:
                names.append(name)
        return ', '.join(names[:4])  # cap at 4 for readability

    for r in rows:
        wk  = r['week']
        day = r['day'].strip().upper()
        if day not in DAYS:
            continue
        day_idx = DAYS.index(day)
        _, date_num = DATE_MAP[(wk, day_idx)]
        if date_num not in meals:
            meals[date_num] = {'lunch':'', 'dinner':''}

    # Group by date_num + period
    by_date_period = defaultdict(list)
    for r in rows:
        period = r.get('period','')
        if period not in ('LUNCH','DINNER'):
            continue
        wk  = r['week']
        day = r['day'].strip().upper()
        if day not in DAYS:
            continue
        day_idx = DAYS.index(day)
        _, date_num = DATE_MAP[(wk, day_idx)]
        by_date_period[(date_num, period)].append(r)

    for (dn, period), items in by_date_period.items():
        if dn not in meals:
            meals[dn] = {'lunch':'', 'dinner':''}
        key = 'lunch' if period == 'LUNCH' else 'dinner'
        meals[dn][key] = extract_meal_name(items)

    # Sort by dateNum order (handle Wk3+ wrapping past 31)
    date_order = [DATE_MAP[(wk,di)][1] for wk in range(1,5) for di in range(7)]
    sorted_meals = sorted(meals.items(), key=lambda x: date_order.index(x[0]) if x[0] in date_order else 99)

    parts = []
    for dn, m in sorted_meals:
        l = js_str(m['lunch'])
        d = js_str(m['dinner'])
        parts.append(f"{dn}:{{lunch:'{l}',dinner:'{d}'}}")
    return 'const MEALS={' + ',\n'.join(parts) + '};'

def _parse_qty_str(row):
    """Extract a short qty string for MOVES display from yield or qty."""
    y = str(row.get('yield', '') or '').strip()
    if y:
        # Clean up: "2u" → "2u", "3 hotels" → "3 hotels", "1/2 hotel" → "0.5u"
        return y.split('/')[0].strip() if '/' in y and 'hotel' not in y.lower() else y
    q = str(row.get('qty', '') or '').strip()
    return q if q else '?'

def _hold_class(hold):
    """Return holdClass for MOVES badge color: g=green ≤2d, y=yellow 3-6d, r=red 7+d."""
    if hold <= 2: return 'g'
    if hold <= 6: return 'y'
    return 'r'

def _abs_day(wk, day_upper):
    """Return absolute day index (0=Wk1 Sun, 27=Wk4 Sat)."""
    if day_upper not in DAYS:
        return None
    return (wk - 1) * 7 + DAYS.index(day_upper)

def _datenum_from_abs(abs_day):
    """Convert absolute day back to (wk, day_idx, dateNum). Returns None if out of range."""
    if abs_day < 0 or abs_day > 27:
        return None
    wk = abs_day // 7 + 1
    di = abs_day % 7
    key = (wk, di)
    if key in DATE_MAP:
        _, dn = DATE_MAP[key]
        return (wk, di, dn)
    return None

def build_moves(rows):
    """Build MOVES JS object from cold-chain items.
    MOVES = { dateNum: [ {dir, item, from, to, qty, run, hold, holdClass, notes, serves/produced} ] }
    Each Rex-bound COOK item generates: SEND (prod_day+1 PM) + PULL (serve_day-1 AM).
    """
    moves = defaultdict(list)  # dateNum → list of move entries

    # Filter to COOK items going to Rex (the primary cold-chain set)
    cold_items = [r for r in rows
                  if r.get('dest_fridge') == 'Rex'
                  and r.get('hold_days', 0) > 0
                  and r.get('type') in ('COOK',)]

    for r in cold_items:
        wk = r['week']
        day_upper = r['day'].strip().upper()
        hold = r['hold_days']
        item = r['item']
        qty = _parse_qty_str(r)
        hc = _hold_class(hold)
        serves_wk = int(r.get('serves_week', 0) or 0)
        serves_day = str(r.get('serves_day', '') or '').strip().upper()

        prod_abs = _abs_day(wk, day_upper)
        if prod_abs is None:
            continue

        # Clean item name for display (strip parentheticals)
        display = re.sub(r'\s*\(.*?\)', '', item).strip()
        display = re.sub(r'^(Cook|Heat|Reheat|Steam|Roast|Bake)\s+', '', display, flags=re.I).strip()

        # SEND: day after production, PM van
        send_abs = prod_abs + 1
        send_info = _datenum_from_abs(send_abs)
        if send_info:
            s_wk, s_di, s_dn = send_info
            s_date = DATE_MAP[(s_wk, s_di)][0]
            serves_str = build_serves(r, wk, DAYS.index(day_upper))
            moves[s_dn].append({
                'dir': 'SEND', 'item': display, 'from': 'Bloor', 'to': 'Rex',
                'qty': qty, 'run': 'PM', 'hold': f'{hold}d', 'holdClass': hc,
                'notes': f'Hold at Rex until pull',
                'serves': serves_str or '',
            })

        # PULL: day before service, AM van
        if serves_wk and serves_day in DAYS:
            serve_abs = _abs_day(serves_wk, serves_day)
            if serve_abs is not None:
                pull_abs = serve_abs - 1
                pull_info = _datenum_from_abs(pull_abs)
                if pull_info:
                    p_wk, p_di, p_dn = pull_info
                    p_date = DATE_MAP[(p_wk, p_di)][0]
                    prod_date = DATE_MAP[(wk, DAYS.index(day_upper))][0]
                    moves[p_dn].append({
                        'dir': 'PULL', 'item': display, 'from': 'Rex',
                        'to': 'Bloor (reheat → Rex hot)',
                        'qty': qty, 'run': 'AM', 'hold': f'{hold}d', 'holdClass': hc,
                        'notes': 'For service',
                        'produced': f'Wk. {wk} {DAY_NAMES[DAYS.index(day_upper)]}',
                    })

    # Build JS
    # Sort by dateNum in calendar order
    date_order = [DATE_MAP[(wk, di)][1] for wk in range(1, 5) for di in range(7)]

    parts = []
    for dn in date_order:
        if dn not in moves:
            continue
        entries = moves[dn]
        entry_strs = []
        for m in entries:
            fields = [f"dir:'{m['dir']}'", f"item:'{js_str(m['item'])}'",
                       f"from:'{m['from']}'", f"to:'{js_str(m['to'])}'",
                       f"qty:'{js_str(m['qty'])}'", f"run:'{m['run']}'",
                       f"hold:'{m['hold']}'", f"holdClass:'{m['holdClass']}'",
                       f"notes:'{js_str(m['notes'])}'"]
            if 'serves' in m and m['serves']:
                fields.append(f"serves:'{js_str(m['serves'])}'")
            if 'produced' in m:
                fields.append(f"produced:'{js_str(m['produced'])}'")
            entry_strs.append('{' + ','.join(fields) + '}')
        parts.append(f"{dn}:[" + ','.join(entry_strs) + ']')

    return 'const MOVES={' + ',\n'.join(parts) + '};'

def build_weeks_obj():
    """Build WEEKS JS object."""
    parts = []
    for wk in range(1,5):
        days_ref = f"...WK{wk}"
        parts.append(
            f"{wk}:{{number:{wk},range:'{WEEK_RANGES[wk]}',"
            f"note:'{js_str(WEEK_NOTE)}',days:[{days_ref}]}}"
        )
    return 'const WEEKS={' + ',\n'.join(parts) + '};'

def extract_block(html, const_name):
    """Extract a const block from hub HTML by brace-counting."""
    start = html.find(f'const {const_name}=')
    if start == -1:
        return ''
    # Find opening brace/bracket
    eq = html.index('=', start) + 1
    while html[eq] in ' \t\n':
        eq += 1
    opener = html[eq]
    closer = ']' if opener == '[' else '}'
    depth  = 0
    i      = eq
    while i < len(html):
        if html[i] == opener:
            depth += 1
        elif html[i] == closer:
            depth -= 1
            if depth == 0:
                # include trailing semicolon if present
                end = i+1
                if end < len(html) and html[end] == ';':
                    end += 1
                return html[start:end]
        i += 1
    return ''

def build_static_fallback(rows):
    """Generate the static fallback HTML for all 4 weeks."""
    # Color maps
    TYPE_COLORS = {
        'SEND AM': ('#c00','#fff'),
        'SEND PM': ('#c55a11','#fff'),
        'COOK':    ('#375623','#fff'),
        'HEAT':    ('#7f6000','#fff'),
        'PREP':    ('#1f3864','#fff'),
        'SOUP':    ('#375623','#fff'),
    }
    SITE_COLORS = {
        'Bloor':'#b4c6e7','LAN':'#fff2cc','GC':'#d9d9d9','Rex':'#c5d9a4'
    }
    SECTION_COLORS = {
        'send-am':'#c00','lunch':'#1f3864','production':'#f2f2f2',
        'dinner':'#375623','send-pm':'#c55a11'
    }
    SECTION_TEXT = {
        'send-am':'#fff','lunch':'#fff','production':'#555',
        'dinner':'#fff','send-pm':'#fff'
    }

    by_day = defaultdict(lambda: defaultdict(list))
    for r in rows:
        wk  = r['week']
        day = r['day'].strip().upper()
        if day not in DAYS:
            continue
        by_day[(wk, day)][r['period']].append(r)

    html_parts = ['<div id="sf">']
    html_parts.append('''
<div id="sf-mob" style="position:sticky;top:0;z-index:100;background:#2e75b6;
color:#fff;padding:10px 16px;font-size:13px;font-family:sans-serif;text-align:center">
📱 For the full interactive version: click <b>Open → Open in app</b><br>
<span style="font-size:11px;opacity:.8">Click 'Trust it' when prompted — the file is from your shared Kitchen folder</span>
</div>''')

    # Title
    html_parts.append('''
<div style="text-align:center;padding:20px 10px 10px;font-family:sans-serif">
<div style="font-size:20px;font-weight:700">CONC Kitchen — Production Hub</div>
<div style="font-size:12px;color:#666;margin-top:4px">May 17 – Jun 13, 2026 · Remediation</div>
</div>''')

    # Week jump links
    jumps = ' '.join(
        f'<a href="#sf-wk{w}" style="background:#2c2a26;color:#fff;padding:8px 16px;'
        f'border-radius:8px;text-decoration:none;font-size:13px;font-family:sans-serif">'
        f'Wk {w}</a>'
        for w in range(1,5)
    )
    html_parts.append(f'<div style="display:flex;gap:8px;justify-content:center;padding:8px 0">{jumps}</div>')

    for wk in range(1, 5):
        html_parts.append(
            f'<div id="sf-wk{wk}" style="background:#2c2a26;color:#fff;font-family:sans-serif;'
            f'font-size:15px;font-weight:700;padding:10px 16px;margin-top:16px;'
            f'scroll-margin-top:60px">Week {wk} — {WEEK_RANGES[wk]}</div>'
        )
        for day_idx, day_upper in enumerate(DAYS):
            date_str, date_num = DATE_MAP[(wk, day_idx)]
            day_name = DAY_NAMES[day_idx]
            day_rows = by_day[(wk, day_upper)]
            if not any(day_rows.values()):
                continue

            # Count summary
            counts = []
            for period, label in [('LUNCH','Lunch'),('SEND AM','AM'),
                                   ('PRODUCTION','Prod'),('DINNER','Dinner'),('SEND PM','PM')]:
                c = len(day_rows.get(period,[]))
                if c:
                    counts.append(f'{label}:{c}')
            summary = ' · '.join(counts) + ' ▸'

            html_parts.append(
                f'<details style="margin:2px 0;font-family:sans-serif">'
                f'<summary style="background:#2c2a26;color:#fff;padding:7px 12px;'
                f'font-size:13px;cursor:pointer;display:flex;justify-content:space-between">'
                f'<span style="font-weight:700">{day_name} {date_str}</span>'
                f'<span style="font-size:11px;opacity:.7">{summary}</span></summary>'
            )

            for period in SECTION_ORDER:
                sec_rows = day_rows.get(period, [])
                if not sec_rows:
                    continue
                sid = SECTION_IDS[period]
                label = SECTION_LABELS[sid]
                bg = SECTION_COLORS.get(sid,'#eee')
                fg = SECTION_TEXT.get(sid,'#000')
                html_parts.append(
                    f'<div style="background:{bg};color:{fg};padding:5px 9px;'
                    f'font-size:10px;text-transform:uppercase;font-weight:700;'
                    f'border-radius:5px;margin:2px 0">{label}</div>'
                )
                for r in sec_rows:
                    typ  = r.get('type','')
                    item = r.get('item','')
                    qty  = str(r.get('qty','') or '')
                    site = r.get('site','') or ''
                    cook = str(r.get('cook_min','') or '')
                    tc, tf = TYPE_COLORS.get(typ, ('#888','#fff'))
                    sc = SITE_COLORS.get(site,'#eee')
                    time_str = f'Overnight' if cook == '480' else (f'{cook} min' if cook else '')
                    detail = ' · '.join(filter(None,[qty, time_str]))
                    serves = build_serves(r, wk, day_idx)
                    serves_span = (f'<span style="background:#fff3cd;color:#856404;'
                                   f'border-radius:3px;padding:1px 4px;font-size:9px;margin-left:4px">'
                                   f'{serves}</span>') if serves else ''
                    html_parts.append(
                        f'<div style="font-size:12px;line-height:1.6;border-left:3px solid #e8e4dd;'
                        f'padding:4px 8px;margin:1px 0">'
                        f'<span style="background:{tc};color:{tf};font-size:8px;text-transform:uppercase;'
                        f'padding:1px 4px;border-radius:3px;margin-right:4px">{typ}</span>'
                        f'<b>{item}</b>'
                        f'<span style="background:{sc};font-size:9px;padding:1px 4px;'
                        f'border-radius:3px;margin-left:4px">{site}</span>'
                        f'{serves_span}'
                        + (f'<div style="font-size:10px;color:#666;font-family:monospace">{detail}</div>' if detail else '')
                        + '</div>'
                    )
            html_parts.append('</details>')
    html_parts.append('</div>')
    return '\n'.join(html_parts)

def generate_hub(rows):
    """Assemble full hub HTML."""
    print("Building hub data blocks...")

    # Load static parts
    shell_html = SHELL_PATH.read_text(encoding='utf-8')
    rcp_block  = RCP_PATH.read_text(encoding='utf-8')
    logic_js   = LOGIC_PATH.read_text(encoding='utf-8')

    # Extract FRIDGE from existing hub (carried forward — item-level detail too complex to auto-generate)
    if HUB_TEMPLATE.exists():
        existing_hub = HUB_TEMPLATE.read_text(encoding='utf-8')
        fridge_block = extract_block(existing_hub, 'FRIDGE')
    else:
        fridge_block = ''
    if not fridge_block:
        print("  ⚠ FRIDGE: no template hub found — using empty block")
        fridge_block = 'const FRIDGE={};'
    else:
        print("  ✓ FRIDGE carried forward from template hub")

    # Generate MOVES from source data (cold-chain items with hold_days > 0)
    print("  Generating MOVES from source data...")
    moves_block = build_moves(rows)
    moves_count = moves_block.count("dir:'SEND'") + moves_block.count("dir:'PULL'")
    send_count = moves_block.count("dir:'SEND'")
    pull_count = moves_block.count("dir:'PULL'")
    print(f"  ✓ MOVES: {send_count} SENDs + {pull_count} PULLs = {moves_count} entries")

    # Generate data blocks
    s_helper = "function S(type,item,qty,site,opts={}){return{type,item,qty,site,...opts}}"
    meals_block = build_meals(rows)
    wk_blocks   = [build_wk_data(rows, wk) for wk in range(1,5)]
    weeks_block = build_weeks_obj()

    # Static fallback
    print("Building static fallback...")
    static_fallback = build_static_fallback(rows)

    # Assemble data block
    data_block = '\n'.join([
        s_helper,
        fridge_block,
        moves_block,
        meals_block,
        *wk_blocks,
        weeks_block,
        rcp_block,
    ])

    # Assemble full HTML
    out = (
        shell_html +
        '\n\n' +
        static_fallback +
        '\n<script>\ndocument.body.classList.add(\'js-ready\');\n' +
        data_block +
        '\n' +
        logic_js +
        '\n</script>\n</body>\n</html>\n'
    )

    # Validate
    print("Validating JS...")
    checks = ['const FRIDGE=','const MOVES=','const MEALS=',
              'const WK1=','const WK2=','const WK3=','const WK4=',
              'const WEEKS=','const RCP=','const RCP_KEYS=',
              'function rcpUrl','function S(']
    for pat in checks:
        count = out.count(pat)
        if count != 1:
            print(f"  ⚠ DUPLICATE or MISSING: '{pat}' appears {count}x")
        else:
            print(f"  ✓ {pat}")

    # Node.js syntax check via temp file
    script_start = out.find('<script>') + 8
    script_end   = out.rfind('</script>')
    js_only      = out[script_start:script_end]
    tmp_js = Path('/tmp/hub_check.js')
    try:
        tmp_js.write_text(js_only)
        result = subprocess.run(['node','--check', str(tmp_js)], capture_output=True)
        if result.returncode == 0:
            print("  ✓ node --check passed")
        else:
            print(f"  ✗ node --check FAILED:\n{result.stderr.decode()[:500]}")
    except FileNotFoundError:
        print("  (node not available — skipping JS syntax check)")

    return out

# ══════════════════════════════════════════════════════════════════════════════
#  PRODUCTION SCHEDULE XLSX GENERATOR  (clean reference output)
# ══════════════════════════════════════════════════════════════════════════════

def generate_prod_schedule(rows):
    """Generate formatted production schedule xlsx matching operational layout."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not available — skipping xlsx generation")
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Type fills for data rows (cols A–J + L)
    TYPE_FILLS = {
        'SEND AM': 'FFFF0000', 'SEND PM': 'FFED7D31',
        'COOK':    'FFE2EFDA', 'HEAT':    'FFFFF2CC',
        'PREP':    'FFDAE3F3', 'SOUP':    'FF92D050',
    }
    # Type font colors (white text for SEND rows)
    TYPE_FONTS = {
        'SEND AM': 'FFFFFFFF', 'SEND PM': 'FFFFFFFF',
    }
    # Site fills for col K (independent of type)
    SITE_FILLS = {
        'Bloor': 'FFB4C6E7', 'LAN': 'FFFFF2CC', 'GC': 'FFD9D9D9',
        'Bloor+LAN': 'FFC5D9A4', 'Bloor/LAN': 'FFC5D9A4', 'Rex': 'FFC5D9A4',
    }
    # Section header config: emoji, label suffix, fill
    SECTION_CFG = {
        'SEND AM':    ('▲', 'SEND AM — Morning van dispatch', 'FFFCE4D6'),
        'LUNCH':      ('☀', 'LUNCH — Items cooked / heated for Lunch service', 'FFDAE3F3'),
        'PRODUCTION': ('⚙', 'PRODUCTION — Advance prep, long cooks, bagging', 'FFF2F2F2'),
        'DINNER':     ('🌙', 'DINNER — Items cooked / heated for Dinner service', 'FFE2EFDA'),
        'SEND PM':    ('▼', 'SEND PM — Evening van dispatch', 'FFED7D31'),
    }
    # Column headers
    COL_HEADERS = ['DAY', 'SECTION', 'TYPE', 'ITEM / TASK', 'QTY', 'PORTION',
                   'PACKAGING', 'YIELD/AMT', 'NOTES / FLAGS', 'COOK TIME', 'SITE',
                   'PROD / SERVICE DATE']
    COL_WIDTHS = [12, 13, 9, 49.3, 14, 10, 12, 15.9, 38, 11, 12, 36]
    NUM_COLS = len(COL_HEADERS)

    for wk in range(1, 5):
        ws = wb.create_sheet(f'Week {wk}  |  {WEEK_RANGES[wk]}')
        wk_rows = [r for r in rows if r['week'] == wk]

        # ── Row 1: Week header ──
        ri = 1
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=NUM_COLS)
        c = ws.cell(ri, 1, f'WEEK {wk}  |  {WEEK_RANGES[wk]}')
        c.font = Font(bold=True, size=13, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF1F3864')
        c.alignment = Alignment(vertical='center')
        ws.row_dimensions[ri].height = 28

        # ── Row 2: Subtitle ──
        ri = 2
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=NUM_COLS)
        c = ws.cell(ri, 1, WEEK_NOTE)
        c.font = Font(size=9, color='FF555555')
        c.fill = PatternFill('solid', fgColor='FFF2F2F2')
        c.alignment = Alignment(vertical='center', wrap_text=True)

        # ── Row 3: Column headers ──
        ri = 3
        for ci, h in enumerate(COL_HEADERS, 1):
            c = ws.cell(ri, ci, h)
            c.font = Font(bold=True, size=9, color='FFFFFFFF')
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(wrap_text=True, vertical='center')
            c.border = border

        ri = 4

        # ── Group rows by day → section ──
        by_day = defaultdict(lambda: defaultdict(list))
        for r in wk_rows:
            day_upper = r['day'].strip().upper()
            if day_upper in DAYS:
                by_day[day_upper][r['period']].append(r)

        for day_idx, day_upper in enumerate(DAYS):
            date_str, _ = DATE_MAP[(wk, day_idx)]
            day_name = DAY_NAMES[day_idx]
            day_sections = by_day[day_upper]
            if not any(day_sections.values()):
                continue

            # ── ◆ Day header ──
            ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=NUM_COLS)
            c = ws.cell(ri, 1, f'◆  {day_upper}  —  {date_str}')
            c.font = Font(bold=True, size=11, color='FFFFFFFF')
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(vertical='center')
            ws.row_dimensions[ri].height = 22
            ri += 1

            for period in SECTION_ORDER:
                sec_rows = day_sections.get(period, [])
                if not sec_rows:
                    continue
                emoji, label_suffix, sec_fill = SECTION_CFG[period]

                # Van load for SEND sections
                van_str = ''
                if period in ('SEND AM', 'SEND PM'):
                    vl = van_load_sum(sec_rows)
                    if vl:
                        van_str = f'  |  VAN LOAD: {vl}'

                # ── Section header ──
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=NUM_COLS)
                c = ws.cell(ri, 1, f'{emoji}  {label_suffix}{van_str}')
                c.font = Font(bold=True, size=9)
                c.fill = PatternFill('solid', fgColor=sec_fill)
                c.alignment = Alignment(vertical='center', wrap_text=True)
                ws.row_dimensions[ri].height = 18
                ri += 1

                # ── Data rows ──
                for row in sec_rows:
                    typ = row.get('type', '')
                    site = row.get('site', '')
                    row_fill = TYPE_FILLS.get(typ, 'FFFFFFFF')
                    font_color = TYPE_FONTS.get(typ, 'FF000000')
                    site_fill = SITE_FILLS.get(site, 'FFB4C6E7')
                    cook = str(row.get('cook_min', '') or '')
                    cook_display = 'Overnight' if cook == '480' else (f'{cook} min' if cook else '')

                    vals = [
                        day_name, '', typ, row['item'],
                        row.get('qty', ''), row.get('portion', ''),
                        row.get('packaging', ''), row.get('yield', ''),
                        row.get('notes', ''), cook_display,
                        site, row.get('prod_date_raw', ''),
                    ]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(ri, ci, v)
                        c.font = Font(size=9, color=font_color)
                        c.alignment = Alignment(wrap_text=True, vertical='center')
                        c.border = border
                        # Col K gets site fill, all others get type fill
                        if ci == 11:  # SITE column
                            c.fill = PatternFill('solid', fgColor=site_fill)
                        else:
                            c.fill = PatternFill('solid', fgColor=row_fill)
                    ri += 1

        # ── Column widths ──
        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    return wb

# ══════════════════════════════════════════════════════════════════════════════
#  DRIVER SCHEDULE XLSX GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def generate_driver_schedule(rows):
    """Generate driver-facing schedule xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin')
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    for wk in range(1,5):
        ws = wb.create_sheet(f'Wk{wk} Driver')

        # Filter to SEND rows only
        send_rows = [r for r in rows
                     if r['week']==wk and r.get('period') in ('SEND AM','SEND PM')]

        headers = ['Day','Date','Van Run','Item','Qty','Route','From→To','Notes']
        for ci, h in enumerate(headers,1):
            c = ws.cell(1,ci,h)
            c.font = Font(bold=True,size=9,color='FFFFFFFF')
            c.fill = PatternFill('solid',fgColor='FF2E75B6')
            c.alignment = Alignment(wrap_text=True,vertical='center')

        # Group by day then AM/PM
        by_day = defaultdict(lambda: {'AM':[],'PM':[]})
        for r in send_rows:
            day  = r['day'].strip().upper()
            run  = r.get('van_run','')
            if day in DAYS:
                by_day[day][run].append(r)

        ri = 2
        for day_upper in DAYS:
            day_idx  = DAYS.index(day_upper)
            date_str,_ = DATE_MAP[(wk, day_idx)]
            for run in ['AM','PM']:
                run_rows = by_day[day_upper][run]
                for r in run_rows:
                    route = build_route(r)
                    frm, to = (route.split('→')[0].strip(), route.split('→')[1].strip()) \
                               if '→' in route else (r.get('site',''), 'Rex')
                    vals = [
                        DAY_NAMES[day_idx], date_str, run,
                        r['item'], r.get('qty',''), route,
                        f'{frm} → {to}', clean_driver_notes(r.get('notes',''), route)
                    ]
                    fill = 'FFFFE699' if run == 'AM' else 'FFFFD966'
                    for ci, v in enumerate(vals,1):
                        c = ws.cell(ri,ci,v)
                        c.font = Font(size=9)
                        c.alignment = Alignment(wrap_text=True,vertical='center')
                        c.fill = PatternFill('solid',fgColor=fill)
                        c.border = border
                    ri += 1

        widths = [12,10,8,45,12,18,18,35]
        for ci,w in enumerate(widths,1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    return wb

# ══════════════════════════════════════════════════════════════════════════════
#  LABOUR REPORT XLSX GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def generate_labour_report(rows):
    """Generate simplified labour/oven time report xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin')
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    # Summary sheet
    ws_sum = wb.create_sheet('Summary')
    sum_headers = ['Week','Day','Date','Site','Bloor Active (min)',
                   'Bloor Oven (min)','LAN Active (min)','LAN Oven (min)',
                   'Total Active (min)','Total Oven (min)','Alert']
    for ci,h in enumerate(sum_headers,1):
        c = ws_sum.cell(1,ci,h)
        c.font = Font(bold=True,size=9,color='FFFFFFFF')
        c.fill = PatternFill('solid',fgColor='FF2E75B6')
        c.alignment = Alignment(wrap_text=True,vertical='center')

    sum_ri = 2
    for wk in range(1,5):
        ws_det = wb.create_sheet(f'Wk{wk} Detail')
        det_headers = ['Day','Type','Item','Site','Cook Min','Active Min','Oven Min','Notes']
        for ci,h in enumerate(det_headers,1):
            c = ws_det.cell(1,ci,h)
            c.font = Font(bold=True,size=9,color='FFFFFFFF')
            c.fill = PatternFill('solid',fgColor='FF375623')
            c.alignment = Alignment(wrap_text=True,vertical='center')

        wk_rows = [r for r in rows
                   if r['week']==wk and r.get('period') not in ('SEND AM','SEND PM')]

        det_ri = 2
        by_day_site = defaultdict(lambda: defaultdict(lambda: {'active':0,'oven':0}))

        for r in wk_rows:
            day = r['day'].strip().upper()
            if day not in DAYS:
                continue
            site = r.get('site','Bloor')
            act, ov = labour_times(r)
            cook = str(r.get('cook_min','') or '')
            # Accumulate for summary
            site_key = 'LAN' if site == 'LAN' else 'Bloor'
            by_day_site[day][site_key]['active'] += act
            by_day_site[day][site_key]['oven']   += ov

            vals = [DAY_NAMES[DAYS.index(day)], r['type'], r['item'],
                    site, cook, act, ov, r.get('notes','')]
            fill = 'FFE2EFDA' if site_key == 'Bloor' else 'FFFFF2CC'
            for ci,v in enumerate(vals,1):
                c = ws_det.cell(det_ri,ci,v)
                c.font = Font(size=9)
                c.alignment = Alignment(wrap_text=True,vertical='center')
                c.fill = PatternFill('solid',fgColor=fill)
                c.border = border
            det_ri += 1

        # Write summary rows for this week
        for day_idx, day_upper in enumerate(DAYS):
            date_str,_ = DATE_MAP[(wk,day_idx)]
            d = by_day_site[day_upper]
            b_act = d['Bloor']['active']
            b_ov  = d['Bloor']['oven']
            l_act = d['LAN']['active']
            l_ov  = d['LAN']['oven']
            t_act = b_act + l_act
            t_ov  = b_ov  + l_ov
            alert = []
            if b_act > 480: alert.append('Bloor active >8h')
            if b_ov  > 360: alert.append('Oven >6h')
            if l_act > 240: alert.append('LAN >4h')
            alert_str = ' | '.join(alert)

            vals = [wk, DAY_NAMES[day_idx], date_str, '',
                    b_act, b_ov, l_act, l_ov, t_act, t_ov, alert_str]
            fill = 'FFFFC7CE' if alert_str else 'FFC6EFCE'
            for ci,v in enumerate(vals,1):
                c = ws_sum.cell(sum_ri,ci,v)
                c.font = Font(size=9)
                c.alignment = Alignment(vertical='center')
                c.fill = PatternFill('solid',fgColor=fill)
                c.border = border
            sum_ri += 1

        # Detail col widths
        for ci,w in enumerate([12,8,45,10,10,12,12,35],1):
            ws_det.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Summary col widths
    for ci,w in enumerate([6,12,10,8,18,18,18,18,18,18,30],1):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Caveat note
    sum_ri += 1
    ws_sum.merge_cells(start_row=sum_ri, start_column=1, end_row=sum_ri, end_column=11)
    c = ws_sum.cell(sum_ri, 1,
        '⚠ UPPER BOUNDS: Oven times assume sequential use. Actual oven time is lower '
        'when items cook in parallel. Alerts flag days that may need attention even after '
        'parallel deductions.')
    c.font = Font(size=9, italic=True, color='FF8B0000')
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws_sum.row_dimensions[sum_ri].height = 36

    return wb

# ══════════════════════════════════════════════════════════════════════════════
#  SOURCE DATA VALIDATION
# ══════════════════════════════════════════════════════════════════════════════

def validate_source(rows):
    """Run data quality checks on source rows. Returns (warnings, errors) lists."""
    warnings = []
    errors = []

    for i, r in enumerate(rows, 1):
        wk   = r.get('week', 0)
        day  = r.get('day', '')
        item = r.get('item', '')
        typ  = r.get('type', '')
        period = r.get('period', '')
        site = r.get('site', '')
        tag  = f"Wk{wk} {day[:3]} R{i}"

        # Missing essentials
        if not item:
            errors.append(f"{tag}: missing item name")
        if not typ:
            errors.append(f"{tag}: missing type — {item[:40]}")
        if not site:
            warnings.append(f"{tag}: missing site — {item[:40]}")

        # SEND rows: hotel_equiv and van_run
        if period in ('SEND AM', 'SEND PM'):
            he = r.get('hotel_equiv', '')
            if he is None or he == '':
                warnings.append(f"{tag}: SEND missing hotel_equiv — {item[:40]}")
            vr = r.get('van_run', '')
            if not vr:
                warnings.append(f"{tag}: SEND missing van_run — {item[:40]}")

        # PRODUCTION rows: serves data
        if period == 'PRODUCTION':
            sw = r.get('serves_week', '')
            sd = r.get('serves_day', '')
            if not sw:
                warnings.append(f"{tag}: PRODUCTION missing serves_week — {item[:40]}")
            if not sd:
                warnings.append(f"{tag}: PRODUCTION missing serves_day — {item[:40]}")

        # Cook time for COOK/HEAT items
        if typ in ('COOK', 'HEAT'):
            cm = str(r.get('cook_min', '') or '').strip()
            if not cm or cm == '0':
                warnings.append(f"{tag}: {typ} missing cook_min — {item[:40]}")

    # dateNum collision check (MEALS key safety)
    datenum_owners = {}  # dateNum → set of weeks that use it
    for r in rows:
        wk = r.get('week', 0)
        day_upper = str(r.get('day', '')).strip().upper()
        if day_upper in DAYS and wk:
            di = DAYS.index(day_upper)
            key = (wk, di)
            if key in DATE_MAP:
                _, dn = DATE_MAP[key]
                datenum_owners.setdefault(dn, set()).add(wk)
    for dn, wks in datenum_owners.items():
        if len(wks) > 1:
            errors.append(f"dateNum COLLISION: dateNum={dn} used by weeks {sorted(wks)} — MEALS keys will overwrite")

    # Summary
    send_count = sum(1 for r in rows if r.get('period') in ('SEND AM', 'SEND PM'))
    send_no_he = sum(1 for r in rows if r.get('period') in ('SEND AM', 'SEND PM')
                     and (r.get('hotel_equiv') is None or r.get('hotel_equiv') == ''))
    cold_chain = sum(1 for r in rows if r.get('hold_days', 0) > 0)

    return warnings, errors, {
        'total_rows': len(rows),
        'send_rows': send_count,
        'send_missing_hotel_equiv': send_no_he,
        'cold_chain_items': cold_chain,
    }

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    hub_only  = '--hub-only'  in sys.argv
    xlsx_only = '--xlsx-only' in sys.argv

    OUTPUT_DIR.mkdir(exist_ok=True)

    src = XLSX_PATH if XLSX_PATH.exists() else CSV_PATH
    print(f"Loading {src}...")
    rows = load_source()
    print(f"  {len(rows)} rows loaded")

    # ── Validate ──────────────────────────────────────
    print("\n── Validation ──────────────────────────────")
    warnings, errors, stats = validate_source(rows)
    print(f"  Rows: {stats['total_rows']}  |  SEND: {stats['send_rows']}  |  Cold-chain: {stats['cold_chain_items']}")
    if stats['send_missing_hotel_equiv']:
        print(f"  ⚠ {stats['send_missing_hotel_equiv']}/{stats['send_rows']} SEND rows missing hotel_equiv")
    if errors:
        print(f"  ✗ {len(errors)} ERROR(S):")
        for e in errors[:10]:
            print(f"      {e}")
        if len(errors) > 10:
            print(f"      ... and {len(errors)-10} more")
    if warnings:
        print(f"  ⚠ {len(warnings)} warning(s) — run with --verbose for details")
        if '--verbose' in sys.argv:
            for w in warnings:
                print(f"      {w}")
    if not errors and not warnings:
        print("  ✓ All checks passed")

    if not xlsx_only:
        print("\n── Hub ──────────────────────────────────────")
        hub_html = generate_hub(rows)
        out_path = OUTPUT_DIR / 'CONC_Production_Hub.html'
        out_path.write_text(hub_html, encoding='utf-8')
        print(f"  Written: {out_path} ({len(hub_html):,} chars)")

    if not hub_only:
        print("\n── Production Schedule ──────────────────────")
        wb = generate_prod_schedule(rows)
        if wb:
            p = OUTPUT_DIR / '01_Production_Schedule_generated.xlsx'
            wb.save(p)
            print(f"  Written: {p}")

        print("\n── Driver Schedule ──────────────────────────")
        wb = generate_driver_schedule(rows)
        if wb:
            p = OUTPUT_DIR / '03_Driver_Schedule_generated.xlsx'
            wb.save(p)
            print(f"  Written: {p}")

        print("\n── Labour Report ────────────────────────────")
        wb = generate_labour_report(rows)
        if wb:
            p = OUTPUT_DIR / 'Labour_Report_generated.xlsx'
            wb.save(p)
            print(f"  Written: {p}")

    print("\n✓ Done.")

if __name__ == '__main__':
    main()
